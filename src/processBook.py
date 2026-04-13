import pyperclip
import subprocess
import os
from PyPDF2 import PdfReader, PdfWriter, PageObject  # Import necessary modules
import shutil
import psutil #to check if Excel is running and terminate it
import re
from openpyxl import load_workbook
from PIL import Image
import io

from extract_pdf_pages import extractSectionPages, yesOrNo
from pdfReverse import reversePages
from toc import clean_leading_dots, correct_line

def print_green(text):
    """Prints text in green color."""
    print("\033[32m" + '\n' + text + "\033[0m")

def print_red(text):
    """Prints text in red color."""
    print("\033[31m" + text + "\033[0m")

def deleteFile(filePath, fileName):
    print("delete file")
    try:
        if os.path.exists(filePath):
            os.remove(filePath)
            print_green(f"{fileName} deleted successfully.\n")
        else:
            print_red(f"\nThe file {filePath} does not exist.")
    except Exception as e:
        print_red(f"\nError deleting the file: {e}")

def moveJpgFile(sourceFolder, destinationFolder):
    while True:
        if not os.path.exists(destinationFolder):
            os.makedirs(destinationFolder)

        for fileName in os.listdir(sourceFolder):
            if re.match(r"^\d{4}\.jpg$", fileName):  # Matches filenames like "1234.jpg"
                sourcePath = os.path.join(sourceFolder, fileName)
                destinationPath = os.path.join(destinationFolder, fileName)
                shutil.move(sourcePath, destinationPath)
                print_green(f"Moved: {fileName} successfully")
                return fileName.removesuffix(".jpg") # Exit the function after successfully moving the file
        else:
            print_red("\njpg file of front cover with danacode file name wasn't found\n")
            choice = yesOrNo("Would you like to try again? (y/n): ")
            
            if not choice:
                print("\nok")
                return None

def checkIfExcelIsOpen():
    """Check if Excel is open and prompt the user to close it."""
    for process in psutil.process_iter(['name', 'pid']):
        if process.info['name'] and 'EXCEL' in process.info['name'].upper():
            # Optionally, print out the process ID for debugging
            print(f"Excel process detected: {process.info['name']} (PID: {process.info['pid']})")
            print_red("Excel is currently open. Please save and close Excel to proceed.")
            return True
    return False

def updateExcelCellAndOpenExcel(danaCode, folderName):
    excelFilePath = r"R:\Documents\001אתר האינטרנט ופרויקטים דיגיטליים\הכנת כתבי עת לאתר\הכנת ספרים לאתר\טבלה מרכזת ספרים דיגיטליים.xlsx"
    
    if danaCode:
        # Check if Excel is open, wait until it's closed
        while checkIfExcelIsOpen():
            input("Press Enter after closing Excel...")  # Wait for user to close Excel
        
        # Load the workbook and select the active worksheet
        workbook = load_workbook(excelFilePath)
        sheet = workbook.active
        
        # Iterate through column B to find the row with the value danaCode
        for row in sheet.iter_rows(min_col=2, max_col=2):  # Only column B
            cell = row[0]  # Get the first (and only) cell in the column
            if cell.value == danaCode:
                # Write folderName in the same row, column F (column 6)
                sheet.cell(row=cell.row, column=6).value = folderName
                print_green(f"Successfully wrote {folderName} in the row of {danaCode}\n")
                break
        
        workbook.save(excelFilePath) # Save the changes to the Excel file
        
        subprocess.run(['start', 'excel', excelFilePath], shell=True) # Open the Excel file
    
def validate_pdf_file(input_pdf_path):
    """Validates the given PDF file path."""
    
    if not os.path.isfile(input_pdf_path):
        return False, "Invalid file path. Please check and try again."
    if not input_pdf_path.lower().endswith('.pdf'):
        return False, "This file is not a PDF."
    return True, None

def getInputPdf():
    input_pdf_path = input("\nEnter the path of the PDF file: ").replace('"','')
    
    while True:
        is_valid, error = validate_pdf_file(input_pdf_path)
        if is_valid:
            break
        print_red(error)
        input_pdf_path = input("\nDrag and drop the PDF file here and press Enter: ").strip('"')
        
    return input_pdf_path

# Function to convert image to PDF page
def image_to_pdf(image_path):
    try:
        img = Image.open(image_path)
        img = img.convert("RGB")  # Ensure the image is in RGB format

        # Save the image as a PDF in memory
        img_bytes = io.BytesIO()
        img.save(img_bytes, format='PDF')
        img_bytes.seek(0)
        
        # Convert to PageObject
        reader = PdfReader(img_bytes)
        return reader.pages[0]
    except Exception as e:
        print_red(f"Error converting image {image_path} to PDF: {e}")
        raise

def addCoverPage(writer, image_path, front = None):
    if front:
        if not os.path.exists(image_path):
            raise FileNotFoundError(f"Front image file not found: {image_path}")
        
        writer.add_page(image_to_pdf(image_path))
        print_green("added front cover")
    
    else:
        if not os.path.exists(image_path):
            raise FileNotFoundError(f"Back image file not found: {image_path}")

        writer.add_page(image_to_pdf(image_path))
        print_green("added back cover")

def extractPages(finFileName, input_pdf_path, folderName):
    if yesOrNo("\nDo you want to extract PDFs for con, pre, chap and eng? (y/n): "):
        thereIsEnglish = extractSectionPages(os.path.join(folderName, finFileName), folderName, True)
        
        if thereIsEnglish == 'y':
            engFile = os.path.join(os.path.dirname(input_pdf_path), f"{folderName}.pdf")
            reversePages(engFile)
            deleteFile(engFile, 'temp english file')

def remove_bleed(file_path, input_folder):
    new_file_name = os.path.join(os.path.dirname(file_path), f"{input_folder}_fin.pdf")
    
    try:
        reader = PdfReader(file_path)
        writer = PdfWriter()        
        
        print("Please enter the amount to crop for each side:")

        def get_crop_value(side):
            while True:
                try:
                    return int(input(f"Enter crop size for {side.upper()}: "))
                except ValueError:
                    print(f"Invalid input. Please enter an integer value for {side.upper()}.")

        crop_left = get_crop_value("left")
        crop_right = get_crop_value("right")
        crop_top = get_crop_value("top")
        crop_bottom = get_crop_value("bottom")
        
        # Add front cover page
        # front_image_path = os.path.join(input_folder, "front.jpg")
        # addCoverPage(writer, front_image_path, True)
        
        print("Please wait, cropping the document...")
        for page in reader.pages:
            page.mediabox.lower_left = (
                page.mediabox.lower_left[0] + crop_left,
                page.mediabox.lower_left[1] + crop_bottom,
            )
            page.mediabox.upper_right = (
                page.mediabox.upper_right[0] - crop_right,
                page.mediabox.upper_right[1] - crop_top,
            )
            writer.add_page(page)
        
        #Add back cover page
        # back_image_path = os.path.join(input_folder, "back.jpg")
        # addCoverPage(writer, back_image_path)        
        
        createToc(reader, writer, file_path)

        with open(new_file_name, "wb") as output_pdf:
            writer.write(output_pdf)

        print_green(f"Finished bleed box removal")
        subprocess.Popen([r"C:\Program Files\Adobe\Acrobat DC\Acrobat\Acrobat.exe", new_file_name])

    except Exception as e:
        print_red(f"Error: An error occurred - {e}")
        
    return new_file_name

def get_pdf_page_count(file_path):
    """Returns the total number of pages in a PDF file."""
    
    try:
        with open(file_path, 'rb') as pdf_file:
            return len(PdfReader(pdf_file).pages)
    except Exception as e:
        print_red(f"Error reading PDF file: {e}")
        return None

def validate_toc_pages(total_pages):
    """
    Prompts the user to enter valid TOC start and end pages.
    
    Args:
        total_pages (int): The total number of pages in the PDF.

    Returns:
        tuple: A tuple containing the valid TOC start and end pages (start_page, end_page).
    """
    while True:
        try:
            toc_start_page = int(input("Enter the starting page number of the TOC: "))
            if 1 <= toc_start_page <= total_pages:
                break
            print_red(f"Page number must be between 1 and {total_pages}.")
        except ValueError:
            print_red("Please enter a valid number.")

    while True:
        try:
            toc_end_page = int(input("Enter the ending page number of the TOC: "))
            if toc_start_page <= toc_end_page <= total_pages:
                print(f"You entered start page: {toc_start_page}, end page: {toc_end_page}")
                confirm = input("Are these correct? (y/n): ").strip().lower()
                if confirm == 'y':
                    return toc_start_page, toc_end_page
                else:
                    print("Re-enter the TOC pages.")
                    return validate_toc_pages(total_pages)
            print_red(f"Page number must be between {toc_start_page} and {total_pages}.")
        except ValueError:
            print_red("Please enter a valid number.")

def askOffset():
    """
    Prompts the user to specify an offset for book page numbering.
    Returns:
        int: The offset entered by the user. Always an integer value (can be positive, negative, or zero).
    """
    while True:
        try:
            user_input = int(input(
                "\nIs there an offset in the book's pages?\n"
                "Please enter the amount of offset pages as a number (positive, negative, or 0 for none): "
            ).strip())
            print(f"You entered offset: {user_input}")
            confirm = input("Is this correct? (y/n): ").strip().lower()
            if confirm == 'y':
                return user_input
            else:
                print("Re-enter the offset.")
        except ValueError:
            print_red("\nInvalid input. Please enter a valid integer.\n")

def extract_toc(pdf_reader, toc_start_page, toc_end_page, total_pages, offset):
    """
    Extracts TOC entries from the specified pages.
    Handles single or two-volume books with different offsets.
    """
    toc = []
    
    print("Extracting toc...")
    
    for page_num in range(toc_start_page - 1, toc_end_page):
        print(f"Page number = {page_num}")
        text = pdf_reader.pages[page_num].extract_text()
        
        if text:
            for line in text.split("\n"):
                print(line)
        else:
            print("No text found on this page.")
        
        # First regex pattern: page number first
        matches = re.findall(r"^\s*(\d+)\s+(.*?)$", text, re.MULTILINE)

        # If no matches, try the second regex pattern: page number last
        if not matches:
            matches = re.findall(r"^(.*?)\s+(\d+)\s*$", text, re.MULTILINE)
            
        for match in matches:
            if len(match) == 2:  # Ensure match has both title and page
                try:
                    if match[0].isdigit():  # First pattern (page number first)
                        page_number = int(match[0]) + offset
                        title = match[1].strip()
                    else:  # Second pattern (title first)
                        page_number = int(match[1]) + offset
                        title = match[0].strip()

                    normalize_toc_title(page_number, title, toc, total_pages)
                except ValueError:
                    print(f"Skipping invalid entry: {match}")

    return toc


def normalize_toc_title(page_number, title, toc, total_pages):
    """
    Cleans and normalizes a table of contents (TOC) title string,
    rearranges parts around delimiters for better readability,
    and appends it with its page number to the TOC list if valid.

    Steps performed:
    1. Removes leading dots or unwanted characters from the title.
    2. Corrects or standardizes the title line format.
    3. Checks for delimiters ("|", "-", " - ") in the title to swap parts around the delimiter
       to improve clarity and presentation.
    4. Removes extra spaces and special spaces.
    5. Appends the processed (title, page_number) tuple to the TOC list if the page number
       is less than the total number of pages.

    Parameters:
    - page_number (int): The page number associated with this TOC entry.
    - title (str): The raw TOC title string to be cleaned and normalized.
    - toc (list): The list where valid TOC entries as tuples (title, page_number) are appended.
    - total_pages (int): The total number of pages in the document, used to validate page numbers.

    Returns:
    None (modifies the toc list in place).

    Example:
    >>> toc = []
    >>> normalize_toc_title(10, "Chapter 1 | Introduction", toc, 100)
    >>> print(toc)
    [('Introduction | Chapter 1', 10)]
    """

    title = clean_leading_dots(title)
    title = correct_line(title)
    # Process the title for better readability
    if "|" in title:
        parts = title.split("|")
        new_title = parts[1].strip() + " | " + parts[0].strip()
    elif "-" in title:
        parts = title.split("-")
        new_title = parts[1].strip() + "-" + parts[0].strip()
    elif " - " in title:
        parts = title.split(" - ")
        new_title = parts[1].strip() + " - " + parts[0].strip()
    else:
        new_title = title.strip()
        new_title = new_title.replace('  ', ' ')
        new_title = new_title.replace(' ', ' ')
    # Append valid TOC entries
    if page_number < total_pages:
        toc.append((new_title, page_number))


def add_bookmarks(pdf_writer, toc, coverOffset):
    """
    Adds bookmarks to the PDF based on the TOC.
    """
    parent_bookmarks = {}
    
    print("Adding bookmarks, please wait...")
    
    for bookmark in toc:
        title, page_num = bookmark
        page_index = page_num + coverOffset
        print(f"{title} at page {page_index}")
        parent_bookmarks[1] = pdf_writer.add_outline_item(title, page_index - 1)
    
    print('\n')

def createToc(reader, pdf_writer, file_path):
    try:
        total_pages = get_pdf_page_count(file_path)
        offset = askOffset()
    
        while True:
            toc_start_page, toc_end_page = validate_toc_pages(total_pages)
            toc = extract_toc(reader, toc_start_page, toc_end_page, total_pages, offset)
            
            if not toc:
                print_red("No TOC detected.\nPerhaps there was some mistake in the input\n")
                while True:
                    choice = input("Would you like to try again? (y/n): ").strip().lower()
                    if choice in {"y", "n"}:
                        break
                    print_red("\nInvalid input. Please enter either 'y' for yes or 'n' for no.\n")
                
                if choice == "n":
                    print("Exiting without processing TOC.")
                    return None
                else:
                    print()
            else:
                print_green("TOC successfully detected")
                break

        add_bookmarks(pdf_writer, toc, offset)  # First offset is used for the base logic
    
    except Exception as e:
        print_red(f"An error occurred: {e}")
        return None
        
def normTitle():
    title = input("Please enter the book's title: ")
    newTitle = title.lower().replace(' ', '_')
    pyperclip.copy(newTitle)  # Copies newTitle to the clipboard
    
def main():
    normTitle()
    
    input_pdf_path = getInputPdf()
        
    sourceFolder = os.path.dirname(input_pdf_path)
    folderName = os.path.basename(sourceFolder)
    
    coversFolder = r"R:\Documents\001אתר האינטרנט ופרויקטים דיגיטליים\הכנת כתבי עת לאתר\הכנת ספרים לאתר\קבצי ספרים מוכנים להעלאה לאמזון\00 תמונות של כריכות ספרים לאמזון"
    
    updateExcelCellAndOpenExcel(moveJpgFile(sourceFolder, coversFolder), folderName)

    finFileName = remove_bleed(input_pdf_path, folderName)
    
    extractPages(finFileName, input_pdf_path, folderName)
    
    # if yesOrNo("Does the pdf file already has bookmarks? (y/n): ") == 'y':
        # fileWithBookmarks = copyToc(fileWithoutBleed, folderName, input_pdf_path)
        
    # else:
        # fileWithBookmarks = processBookNoToc(fileWithoutBleed, folderName)
        
    # deleteFile(fileWithoutBleed, "no bleed box file")

    # if fileWithBookmarks is None or fileWithBookmarks == "":
        # print("Error: fileWithBookmarks is None or empty")

    # fileWithCover = addCoverAndBlankPages(os.path.abspath(fileWithBookmarks), folderName)
    
    # if fileWithCover:  # Check if the function returned a path (not None)
        # deleteFile(fileWithBookmarks, "temp toc file")
        # subprocess.Popen([r"C:\Program Files\Adobe\Acrobat DC\Acrobat\Acrobat.exe", fileWithCover])
    # else:
        # print_red("Error: Could not create the PDF with cover and blank pages.")

if __name__ == "__main__":
    main()