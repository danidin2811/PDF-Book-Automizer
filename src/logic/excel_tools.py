import os
import subprocess
import logging
from pathlib import Path
from typing import Tuple, Optional

import psutil
from openpyxl import load_workbook

from src.constants import BOOK_TRACKER_EXCEL_FILE_PATH, FOLDER_NAME_COL
from src.logic.file_operations import validate_csv_path
from src.logic.interface_controller import AppInterface
from utils.input_output_tools import print_red, print_green
import csv


logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')


def col_letter_to_index(letter: str) -> int:
    """
    Converts Excel column letter to a 1-based index (A=1, B=2, L=12).
    """
    index = 0
    for char in letter.upper():
        # ord('A') is 65. So 'A' - 64 = 1.
        index = index * 26 + (ord(char) - 64)
    return index


def get_new_toc_entries(csv_path):
    """Extracts TOC data with detailed error logging."""
    new_entries = []

    if not os.path.exists(csv_path):
        print(f"DEBUG: File not found at {csv_path}")
        return []

    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        # Use DictReader to handle header variations
        reader = csv.DictReader(f)

        for line_num, row in enumerate(reader, start=2):  # Header is line 1
            try:
                # 1. Level Check
                level_raw = row.get("level", "").strip()
                if not level_raw:
                    raise ValueError("Missing 'level' field")
                level = int(level_raw)

                # 2. Title Check
                title = row.get("title", "").strip()
                if not title:
                    raise ValueError("Missing 'title' field")

                # 3. Page Number Check (Handling both 'page_number' and 'page number')
                page_val = row.get("page_number") or row.get("page number")

                # Allow empty page numbers for level 1 (sections/parts)
                if not page_val or not page_val.strip():
                    if level == 1:
                        page = 0
                    else:
                        raise ValueError(f"Missing page number for level {level} entry")
                else:
                    # Clean the page string in case there are non-digit characters
                    clean_page = ''.join(c for c in page_val if c.isdigit())
                    if not clean_page:
                        raise ValueError(f"Invalid page format: '{page_val}'")
                    page = int(clean_page)

                new_entries.append({
                    "level": level,
                    "title": title,
                    "page": page
                })

            except ValueError as ve:
                print(f"Row {line_num} Data Error: {ve} | Content: {row}")
            except Exception as e:
                print(f"Row {line_num} Unexpected Error: {e}")

    return new_entries


def process_toc_extraction(initial_csv_path):
    current_path = initial_csv_path

    while True:
        is_valid, result_or_error = validate_csv_path(current_path)

        if not is_valid:
            print(f"\n[!] Path Error: {result_or_error}")
            current_path = input("Please enter the correct path to the .csv file: ")
            continue  # Go back to start of loop to check the new path

        # Step 2: Try to extract entries
        # result_or_error is now the cleaned path string
        entries = get_new_toc_entries(result_or_error)

        if entries:
            print("-" * 30)
            print(f"✅ Success! Loaded {len(entries)} entries.")
            print("-" * 30)
            return entries

        # Step 3: Handle empty/invalid file content
        print("\n" + "!" * 40)
        print("[!] CRITICAL: No valid entries found in the file.")
        print(f"File Path: {result_or_error}")
        print("Ensure headers are exactly: level, title, page_number")
        print("!" * 40 + "\n")

        choice = input("Would you like to try again? (y = retry file, p = change path, n = exit): ").lower()

        if choice == 'p':
            current_path = input("Enter new file path: ")
        elif choice != 'y':
            print("Exiting TOC extraction.")
            return None


def get_lock_status(filepath: Path) -> str:
    """
    Determines the specific nature of a file lock.

    Returns:
        str: 'local' if Excel is running on this machine,
             'remote' if the file is locked but Excel is not running locally,
             'none' if the file is accessible.
    """

    # 1. Check if Excel is running locally
    local_excel_active = False
    for process in psutil.process_iter(['name']):
        try:
            if process.info['name'] and 'EXCEL' in process.info['name'].upper():
                local_excel_active = True
                break
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            continue

    # 2. Check if the file itself is locked
    file_is_accessible = True
    if filepath.exists():
        try:
            # Atomic rename check for network/local locks
            os.rename(str(filepath), str(filepath))
        except (OSError, PermissionError):
            file_is_accessible = False

    # 3. Distinguish the cases
    if not file_is_accessible:
        return 'local' if local_excel_active else 'remote'

    return 'none'


def update_excel_cell(row_index: int, col_index: int, sheet_name: str, value: str, interface:AppInterface):
    """
    Directly updates a specific cell using a pre-found row index.
    """
    try:
        workbook = load_workbook(BOOK_TRACKER_EXCEL_FILE_PATH, data_only=False)
        sheet = workbook[sheet_name]

        sheet.cell(row=row_index, column=col_index).value = value

        workbook.save(BOOK_TRACKER_EXCEL_FILE_PATH)
        interface.print_success(f"Successfully wrote {value} to {row_index} x {col_index} in {sheet_name} sheet")
        return True
    except PermissionError:
        print_red("Cannot save: Please close the Excel file!")
        return False


def find_danacode_row(dana_code: str) -> Tuple[bool, Optional[int]]:
    """
    Searches for the DanaCode in Column B and returns (Success, Row_Index).
    """

    if not dana_code:
        return False, None

    try:
        # Load workbook (ReadOnly=True makes it faster if we are just searching)
        workbook = load_workbook(BOOK_TRACKER_EXCEL_FILE_PATH, data_only=False)
        sheet = workbook["ראשי"]
        target = str(dana_code).strip()

        for row in sheet.iter_rows(min_col=2, max_col=2):
            cell = row[0]
            if cell.value and str(cell.value).strip() == target:
                return True, cell.row

        logging.warning(f"DanaCode {target} not found in Column B.")
        return False, None

    except PermissionError:
        print_red(f"Access Denied: {BOOK_TRACKER_EXCEL_FILE_PATH.name} is open.")
        return False, None
    except Exception as e:
        logging.error(f"Excel Error: {e}")
        return False, None


def open_tracker_in_excel() -> None:
    """Opens the tracker using the system default handler for Excel files."""
    try:
        # Use str() for subprocess compatibility
        subprocess.run(['start', 'excel', str(BOOK_TRACKER_EXCEL_FILE_PATH)], shell=True, check=True)
    except subprocess.CalledProcessError as e:
        logging.error(f"Failed to open Excel: {e}")


def run_excel_update_workflow(row_index: int, folder_name: str) -> bool:
    """
    Orchestrates the Excel update with specific feedback on lock types.
    """
    while True:
        status = get_lock_status(BOOK_TRACKER_EXCEL_FILE_PATH)

        if status == 'none':
            break

        if status == 'local':
            print_red("The tracker is open on your computer.")
            print("Please save and close your local Excel window.")

        else:
            print_red("The tracker is locked by another user or another computer.")
            print("Please wait for them to finish or ask them to close the file.")

        user_choice = input("\nPress Enter to retry, or type 'c' to cancel: ").strip().lower()
        if user_choice == 'c':
            return False  # Properly returns bool

    # Proceed with the update
    update_excel_cell(row_index,col_letter_to_index(FOLDER_NAME_COL),"ראשי",folder_name)

    open_tracker_in_excel()
    return True


def get_password_from_excel(row_index: int) -> Optional[str]:
    """
    Directly retrieves a password from a known row index using openpyxl.
    """
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(BOOK_TRACKER_EXCEL_FILE_PATH, data_only=True)
        sheet = workbook["ראשי"]

        # Column 13 is 'M' (where the password usually lives)
        password = sheet.cell(row=row_index, column=13).value

        return str(password) if password else None

    except Exception as e:
        print(f"Error retrieving password: {e}")
        return None