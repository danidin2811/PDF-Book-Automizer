import pyperclip
import subprocess
import os
from PyPDF2 import PdfReader, PdfWriter, PageObject
import psutil #to check if Excel is running and terminate it
import re
from openpyxl import load_workbook
from PIL import Image
# for spinner
import sys
import time
import threading

from extract_pdf_pages import extractSectionPages, yesOrNo, debugPrint
from pdfReverse import reversePages
from toc import clean_leading_dots, correct_line
from processBook import print_green, print_red, deleteFile, moveJpgFile, getInputPdf, image_to_pdf, get_pdf_page_count, validate_toc_pages, askOffset

spinner_running = False  # Global flag to control the spinner


def spinner():
    while spinner_running:
        for char in "-/|\\":
            sys.stdout.write(f"\rWriting PDF... {char}")
            sys.stdout.flush()
            time.sleep(0.1)

def checkIfExcelIsOpen():
    """Check if Excel is open and prompt the user to close it."""
    for process in psutil.process_iter(['name', 'pid']):
        info = process.as_dict(attrs=['name', 'pid'])
        if info['name'] and 'EXCEL' in info['name'].upper():
            print(f"Excel process detected: {info['name']} (PID: {info['pid']})")
            print_red("Excel is currently open. Please save and close Excel to proceed.")
            return True
    return False

def updateExcelCellAndOpenExcel(danaCode, folderName):
    excelFilePath = r"R:\Documents\001אתר האינטרנט ופרויקטים דיגיטליים\הכנת כתבי עת לאתר\הכנת ספרים לאתר\טבלה מרכזת ספרים דיגיטליים.xlsx"
    
    if danaCode:
        # Check if Excel is open, wait until it's closed
        while checkIfExcelIsOpen():
            input("Press Enter after closing Excel...")  # Wait for user to close Excel
        
        # Load the workbook and select the active worksheet
        workbook = load_workbook(excelFilePath)
        
        sheet_name = "ראשי"
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            print_red(f"Sheet '{sheet_name}' not found in workbook.")
            return
        
        selectedRow = None
        
        # Iterate through column B to find the row with the value danaCode
        for row in sheet.iter_rows(min_col=2, max_col=2):  # Only column B
            cell = row[0]  # Get the first (and only) cell in the column

            if str(cell.value).strip() == str(danaCode).strip():
                # Write folderName in the same row, column G (column 7)
                sheet.cell(row=cell.row, column=12).value = folderName
                selectedRow = cell.row
                print_green(f"Successfully wrote {folderName} in the row of {danaCode}\n")
                break                
        
        if selectedRow is None:
            print_red(f"Couldn't find row with danacode {danaCode}")
            
        workbook.save(excelFilePath) # Save the changes to the Excel file
        
        subprocess.run(['start', 'excel', excelFilePath], shell=True) # Open the Excel file

def addCoverPage(writer, coverPath, blankPage, front=None, total_pages=None):
    def ask_to_retry_or_skip(coverType):
        while True:
            choice = input(f"\n{coverType} cover image not found at '{coverPath}'.\nDo you want to provide a new path? (y/n): ").strip().lower()
            if choice == 'y':
                new_path = input(f"Enter the correct path for the {coverType} cover image: ").strip('"')
                if os.path.exists(new_path):
                    return new_path
                else:
                    print_red(f"File not found at '{new_path}'. Try again.")
            elif choice == 'n':
                print(f"\nSkipping {coverType} cover.\n")
                return None
            else:
                print_red("\nInvalid choice. Please enter 'y' or 'n'.")

    coverType = "Front" if front else "Back"

    while not os.path.exists(coverPath):
        coverPath = ask_to_retry_or_skip(coverType)
        if coverPath is None:
            return  # Skip adding this cover and return immediately

    if front:
        writer.add_page(image_to_pdf(coverPath))
        writer.add_page(blankPage)
        print_green(f"Added front cover")

    else:
        if total_pages % 2 == 0:
            writer.add_page(blankPage)
            print_green(f"Added back blank page")
        writer.add_page(image_to_pdf(coverPath))
        print_green(f"Added back cover")


def extractPages(finFileName, input_pdf_path, folderName):
    if yesOrNo("\nDo you want to extract PDFs for con, pre, chap and eng? (y/n): "):
        if yesOrNo("\nDoes the book have front cover? (y/n): "):
            withCover = True
        else:
            withCover = False
            
        thereIsEnglish = extractSectionPages(os.path.join(folderName, finFileName), folderName, withCover)

        if thereIsEnglish:
            engFile = os.path.join(os.path.dirname(input_pdf_path), f"{folderName}.pdf")
            reversePages(engFile)
            deleteFile(engFile, 'temp english file')

def get_integer_input(prompt):
    while True:
        try:
            return int(input(prompt))
        except ValueError:
            print_red("Invalid input. Please enter an integer.")

def get_crop_choice():
    print("\nChoose cropping method:")
    print("1. Set new page dimensions (center crop)")
    print("2. Crop by specifying points from each side")
    print("3. No crop")
    
    while True:
        choice = input("Enter 1, 2, or 3: ").strip()
        if choice in ("1", "2", "3"):
            return choice
        print_red("Invalid choice. Please enter 1, 2, or 3.")

def crop_page_center(page, target_width, target_height):
    orig_width = page.mediabox.width
    orig_height = page.mediabox.height
    margin_x = round((orig_width - target_width) / 2)
    margin_y = round((orig_height - target_height) / 2)

    page.mediabox.lower_left = (margin_x, margin_y)
    page.mediabox.upper_right = (
        round(orig_width - margin_x), 
        round(orig_height - margin_y)
    )

def crop_page_sides(page, left, right, top, bottom):
    page.mediabox.lower_left = (
        page.mediabox.lower_left[0] + left,
        page.mediabox.lower_left[1] + bottom
    )
    page.mediabox.upper_right = (
        page.mediabox.upper_right[0] - right,
        page.mediabox.upper_right[1] - top
    )

def remove_bleed(file_path, input_folder):
    new_file_name = os.path.join(os.path.dirname(file_path), f"{input_folder}_fin.pdf")

    try:
        reader = PdfReader(file_path)
        total_pages = len(reader.pages)
        writer = PdfWriter()

        cropChoice = get_crop_choice()

        if cropChoice == "1":
            while True:
                new_width = get_integer_input("Enter the new WIDTH: ")
                new_height = get_integer_input("Enter the new HEIGHT: ")

                orig_width = reader.pages[0].mediabox.width
                orig_height = reader.pages[0].mediabox.height

                if new_width > orig_width or new_height > orig_height:
                    print_red(f"New dimensions {new_width}x{new_height} exceed original {orig_width}x{orig_height}. Try smaller values.")
                else:
                    break

            blank_page = PageObject.create_blank_page(writer, new_width, new_height)

        elif cropChoice == "2":
            crop_left = get_integer_input("Enter crop size for LEFT: ")
            crop_right = get_integer_input("Enter crop size for RIGHT: ")
            crop_top = get_integer_input("Enter crop size for TOP: ")
            crop_bottom = get_integer_input("Enter crop size for BOTTOM: ")

            sample_page = reader.pages[0]
            width = sample_page.mediabox.width - crop_left - crop_right
            height = sample_page.mediabox.height - crop_top - crop_bottom
            blank_page = PageObject.create_blank_page(writer, width, height)
            
        else:
            sample_page = reader.pages[0]
            blank_page = PageObject.create_blank_page(writer, sample_page.mediabox.width, sample_page.mediabox.height)

        # Add front cover page
        frontCoverPath = os.path.join(os.path.dirname(file_path), "front.jpg")
        addCoverPage(writer, frontCoverPath, blank_page, True, total_pages)

        print("Please wait, cropping the document...")

        for page in reader.pages:
            if cropChoice == "1":
                crop_page_center(page, new_width, new_height)
            elif cropChoice == "2":
                crop_page_sides(page, crop_left, crop_right, crop_top, crop_bottom)
            # No crop needed for choice "3"

            writer.add_page(page)

        # Add back cover page
        backCoverPath = os.path.join(os.path.dirname(file_path), "back.jpg")
        addCoverPage(writer, backCoverPath, blank_page, False, total_pages)

        has_cover = bool(frontCoverPath and backCoverPath)
        createToc(reader, writer, file_path, has_cover)

        global spinner_running
        spinner_running = True
        spinner_thread = threading.Thread(target=spinner)
        spinner_thread.start()

        try:
            with open(new_file_name, "wb") as output_pdf:
                writer.write(output_pdf)
        finally:
            spinner_running = False
            spinner_thread.join()
            print_green("Finished bleed box removal")

        subprocess.Popen([r"C:\Program Files\Adobe\Acrobat DC\Acrobat\Acrobat.exe", new_file_name])

    except Exception as e:
        print_red(f"Error: An error occurred - {e}")

    return new_file_name

def extract_toc(pdf_reader, toc_start_page, toc_end_page, total_pages):
    """
    Extracts TOC entries from the specified pages.
    Handles single or two-volume books with different offsets.
    """
    toc = []
    
    print("Extracting toc...")
    
    for page_num in range(toc_start_page - 1, toc_end_page):
        print(f"Page number = {page_num}")
        text = pdf_reader.pages[page_num].extract_text()

        # First regex pattern: page number first
        matches = re.findall(r"^\s*(\d+)\s+(.*?)$", text, re.MULTILINE)

        # If no matches, try the second regex pattern: page number last
        if not matches:
            matches = re.findall(r"^(.*?)\s+(\d+)\s*$", text, re.MULTILINE)
            
        for match in matches:
            if len(match) == 2:  # Ensure match has both title and page
                try:
                    if match[0].isdigit():  # First pattern (page number first)
                        page_number = int(match[0])
                        title = match[1].strip()
                    else:  # Second pattern (title first)
                        page_number = int(match[1])
                        title = match[0].strip()

                    title = clean_leading_dots(title)
                    title = correct_line(title)

                    # Process the title for better readability

                    if "|" in title:
                        parts = title.split("|")
                        new_title = parts[1].strip() + " | " + parts[0].strip()
                    elif "-" in title:
                        parts = title.split("-")
                        new_title = parts[1].strip() + "-" + parts[0].strip()
                    elif " - " in title:
                        parts = title.split(" - ")
                        new_title = parts[1].strip() + " - " + parts[0].strip()
                    else:
                        new_title = title.strip()
                        new_title = new_title.replace('  ', ' ')
                        new_title = new_title.replace(' ', ' ')

                    # Append valid TOC entries
                    if page_number < total_pages:
                        toc.append((new_title, page_number))
                except ValueError:
                    print(f"Skipping invalid entry: {match}")

    return toc

def add_bookmarks(pdf_writer, toc, coverOffset, hasCover):
    """
    Adds bookmarks to the PDF based on the TOC.
    """
    parent_bookmarks = {}
    if hasCover:
        coverOffset += 2
    
    print("Adding bookmarks, please wait...")

    for bookmark in toc:
        title, page_num = bookmark
        page_index = page_num + coverOffset
        print(f"{title} at page {page_index}")
        parent_bookmarks[1] = pdf_writer.add_outline_item(title, page_index - 1)
    
    print_green('Finished adding bookmarks\n')

def createToc(reader, pdf_writer, file_path, hasCover):
    try:
        total_pages = get_pdf_page_count(file_path)
        offset = askOffset()

        if not yesOrNo("Does the book have a table of contents? (y/n): "):
            print("Skipping TOC processing.")
            return None

        toc = None
        while True:
            toc_start_page, toc_end_page = validate_toc_pages(total_pages)
            toc = extract_toc(reader, toc_start_page, toc_end_page, total_pages)

            if toc:
                print_green("TOC successfully detected")
                break  # Exit retry loop

            print_red("No TOC detected.\nPerhaps there was some mistake in the input\n")
            retry = yesOrNo("Would you like to try again? (y/n): ").strip().lower()
            if not retry:
                print("Exiting without processing TOC.")
                return None
            print()  # Blank line before retry

        add_bookmarks(pdf_writer, toc, offset, hasCover)

    except Exception as e:
        print_red(f"An error occurred: {e}")
        return None

def format_book_title(title: str) -> str:

    small_words = {'and', 'or', 'the', 'of', 'in', 'on', 'a', 'an'}
    words = title.split()
    formatted_words = [
        word.capitalize() if word.lower() not in small_words or i == 0 else word.lower()
        for i, word in enumerate(words)
    ]
    return ' '.join(formatted_words)

def normTitle():
    title = input("Enter book title: ")

    # Replace hyphens with spaces
    title = title.replace('-', ' ')

    formattedBookTitle = format_book_title(title)
    print(f"Book title is: {formattedBookTitle}")

    # Remove non-alphanumeric characters except spaces
    newTitle = re.sub(r'[^a-zA-Z0-9\s]', '', title)

    # Replace multiple spaces with one, convert to lowercase, and replace spaces with underscores
    newTitle = re.sub(r'\s+', ' ', newTitle).strip().lower().replace(' ', '_')

    print(f"Normalized title: {newTitle}")

    
def main():
    normTitle()

    input_pdf_path = getInputPdf()

    sourceFolder = os.path.dirname(input_pdf_path)
    folderName = os.path.basename(sourceFolder)
    
    coversFolder = r"R:\Documents\001אתר האינטרנט ופרויקטים דיגיטליים\הכנת כתבי עת לאתר\הכנת ספרים לאתר\קבצי ספרים מוכנים להעלאה לאמזון\00 תמונות של כריכות ספרים לאמזון"
    
    updateExcelCellAndOpenExcel(moveJpgFile(sourceFolder, coversFolder), folderName)

    finFileName = remove_bleed(input_pdf_path, folderName)

    extractPages(finFileName, input_pdf_path, folderName)

if __name__ == "__main__":
    main()